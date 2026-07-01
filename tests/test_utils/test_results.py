import csv

import openpyxl

from hpm.utils.results import build_report

_COLS = [
    "epoch", "train/loss", "train/id", "val/top1", "val/top5",
    "val/f1_macro", "val/f1_weighted", "val/precision_macro",
    "val/recall_macro", "val/balanced_acc", "lr", "time_s",
]


def _write_run(runs_dir, name, mode, seed, top1_by_epoch):
    """Write a minimal results.csv mirroring train_hpm.py's per-epoch schema."""
    d = runs_dir / f"{name}_{mode}_seed{seed}"
    d.mkdir(parents=True, exist_ok=True)
    with open(d / "results.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=_COLS)
        w.writeheader()
        for epoch, top1 in enumerate(top1_by_epoch):
            w.writerow({c: 0.0 for c in _COLS} | {
                "epoch": epoch, "val/top1": top1, "val/f1_macro": top1 - 0.05,
            })


def test_build_report_aggregates_seeds(tmp_path):
    runs = tmp_path / "runs"
    # HPM clearly beats the baselines; best epoch is the last (monotonic top1).
    per_model = {
        "cnn": {42: [0.50, 0.55], 43: [0.52, 0.57]},
        "vit": {42: [0.58, 0.62], 43: [0.60, 0.64]},
        "hpm": {42: [0.70, 0.75], 43: [0.72, 0.77]},
    }
    modes, seeds = ["cnn", "vit", "hpm"], [42, 43]
    for mode, by_seed in per_model.items():
        for seed, curve in by_seed.items():
            _write_run(runs, "exp", mode, seed, curve)

    out = build_report("exp", modes, seeds, runs_dir=runs)
    assert out is not None and out.exists()

    wb = openpyxl.load_workbook(out)
    # One sheet per run + the three summary sheets.
    assert set(wb.sheetnames) >= {"ModelSummary", "BestPerRun", "Comparison"}
    assert "hpm_s42" in wb.sheetnames

    # ModelSummary: best-epoch top1 averaged across seeds → hpm = mean(0.75, 0.77).
    summary = {r[0]: r for r in wb["ModelSummary"].iter_rows(values_only=True)}
    header = summary["model"]
    mean_col = header.index("top1_mean")
    n_col = header.index("n_seeds")
    assert summary["hpm"][n_col] == 2
    assert abs(summary["hpm"][mean_col] - 0.76) < 1e-6

    # Comparison sheet reports an HPM-vs-baseline delta > 0 for both baselines.
    cmp_rows = list(wb["Comparison"].iter_rows(values_only=True))
    cmp_header = cmp_rows[0]
    delta_i = cmp_header.index("delta")
    top1_deltas = [
        r[delta_i] for r in cmp_rows[1:]
        if r[cmp_header.index("metric")] == "top1"
    ]
    assert top1_deltas and all(d > 0 for d in top1_deltas)


def test_build_report_no_runs_returns_none(tmp_path):
    assert build_report("missing", ["cnn"], [42], runs_dir=tmp_path) is None
